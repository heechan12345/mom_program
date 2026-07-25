"""가족 입출금 내역 정리 프로그램

'SMS Backup & Restore' 등으로 내보낸 문자 백업 XML 파일(농협/우리은행/하나은행
입출금 문자 포함)을 읽어서, 지정한 연/월 범위의 거래만 골라 엑셀로 정리한다.
"""
import os
import re
import json
import datetime as dt
import threading
import traceback
import xml.etree.ElementTree as ET

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext

from openpyxl import Workbook, load_workbook


# ----------------------------------------------------------------------
# 파싱 규칙
#
# 문자 원문 예시 (아래는 형식 설명을 위한 가상의 예시 데이터임):
#   우리 07/24 15:56 / *123456 / 입금 100,000원 / (주)예시상사 / 잔액 1,000,000원
#   하나 07/24 18:09 / 111******22333 / 출금 50,000원 / 예시결제 / 잔액 2,000,000원
#   농협 입금100,000원 / 07/23 11:15 111-****-2222-33 홍길동 잔액300,000원   (일반)
#   농협04/22 20:24 111-****-2222-33 자동출금10,000원(적요) 잔액40,000원      (자동출금, 줄바꿈 없음)
# ----------------------------------------------------------------------

BANK_NAMES = {'우리': '우리은행', '하나': '하나은행', '농협': '농협'}
BANK_ORDER = ['농협', '우리은행', '하나은행']
KST = dt.timezone(dt.timedelta(hours=9))

ACCOUNT_RE = re.compile(
    r'(?:\*\d{5,8}|\d{2,3}\*{3,8}\d{3,6}|\d{2,4}-\*{2,6}-\d{2,6}-\d{1,4})'
)
DATE_RE = re.compile(r'(\d{1,2})/(\d{1,2})')
AMOUNT_TOKEN = r'(?:[\d,]|기)+'
# "자동출금"처럼 입금/출금 앞에 접두어가 붙는 경우도 있어 접두어까지 통째로 지운다.
TXN_RE = re.compile(r'(?:자동)?(입금|출금)\s*(' + AMOUNT_TOKEN + r')\s*원')
BALANCE_RE = re.compile(r'잔액\s*(' + AMOUNT_TOKEN + r')\s*원')

ANCHOR_RE = re.compile(
    r'(?P<bank>우리|하나|농협)(?=\s*\d{1,2}/\d{1,2})'
    r'|(?P<bank2>농협)(?=\s*(?:입금|출금)\s*(?:[\d,]|기)+\s*원)'
)

TOKEN_BLOCKLIST_RE = re.compile(
    r'확인|발신|번호|바로가기|거래내역|^알림$|^오전$|^오후$|^입금$|^출금$|^잔액|요일$|^\d+월$|^\d+일$'
)


def normalize_amount(raw: str) -> int:
    """OCR이 '2,'를 '기'로 잘못 읽는 현상을 보정하여 정수 금액으로 변환.

    (스크린샷 OCR 시절의 잔재지만 문자 원문에는 '기'가 나올 일이 없으므로
    두어도 해가 없다.)
    """
    fixed = raw.replace('기', '2,').replace(',', '')
    return int(fixed)


def extract_time(block: str, start: int):
    """날짜 매칭 바로 뒤 구간에서 시간을 찾는다. 못 찾으면 None."""
    window = block[start:start + 14]
    m = re.match(r'\s*(\d)\s*(\d)\s*[:•.]\s*(\d)\s*(\d)', window)
    if m:
        return f"{m.group(1)}{m.group(2)}:{m.group(3)}{m.group(4)}"
    m = re.match(r'\s*(\d{2})(\d{2})(?!\d)', window)
    if m:
        return f"{m.group(1)}:{m.group(2)}"
    return None


def extract_description(remainder: str) -> str:
    for tok in remainder.split():
        if TOKEN_BLOCKLIST_RE.search(tok):
            continue
        letters = re.sub(r'[^가-힣A-Za-z]', '', tok)
        if len(letters) < 2:
            continue
        return tok
    return ''


def _find_text_field(node):
    """RCS(리치카드) JSON 구조 안에서 실제 문자 내용이 담긴 'text' 값을 재귀적으로 찾는다."""
    if isinstance(node, dict):
        val = node.get('text')
        if isinstance(val, str) and any(b in val for b in ('우리', '하나', '농협')):
            return val
        for v in node.values():
            found = _find_text_field(v)
            if found:
                return found
    elif isinstance(node, list):
        for item in node:
            found = _find_text_field(item)
            if found:
                return found
    return None


def extract_body_text(raw_body: str) -> str:
    """일반 SMS는 그대로, RCS 리치카드(JSON)는 안의 실제 문자 텍스트를 꺼낸다."""
    stripped = raw_body.strip()
    if not stripped.startswith('{'):
        return raw_body
    try:
        data = json.loads(raw_body)
    except (json.JSONDecodeError, ValueError):
        return raw_body
    found = _find_text_field(data)
    return found if found else raw_body


def split_messages(text: str):
    """텍스트를 은행 토큰 기준 메시지 블록들로 분리한다."""
    matches = list(ANCHOR_RE.finditer(text))
    blocks = []
    for i, m in enumerate(matches):
        bank_token = m.group('bank') or m.group('bank2')
        start = m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        blocks.append((bank_token, start, text[start:end]))
    return blocks


def extract_fields(bank_token: str, block: str):
    """블록 하나에서 날짜(월/일)를 뺀 나머지 필드를 뽑는다.

    성공 시 (dict, None), 실패 시 (None, 사유).
    """
    bank = BANK_NAMES[bank_token]

    dm = DATE_RE.search(block)
    if not dm:
        return None, '날짜를 찾지 못함'

    txn = TXN_RE.search(block)
    if not txn:
        return None, '입금/출금 금액을 찾지 못함'

    bal = BALANCE_RE.search(block)
    if not bal:
        return None, '잔액을 찾지 못함'

    try:
        amount = normalize_amount(txn.group(2))
        balance = normalize_amount(bal.group(1))
    except ValueError:
        return None, '금액 숫자 변환 실패'

    time_str = extract_time(block, dm.end())
    acct_m = ACCOUNT_RE.search(block)
    account = acct_m.group(0) if acct_m else ''

    chars = list(block)

    def blank(span):
        s, e = span
        for i in range(s, e):
            chars[i] = ' '

    blank((0, len(bank_token)))
    blank(dm.span())
    blank(txn.span())
    blank(bal.span())
    if acct_m:
        blank(acct_m.span())
    remainder = ''.join(chars)
    desc = extract_description(remainder)

    return {
        '은행': bank,
        'month': int(dm.group(1)),
        'day': int(dm.group(2)),
        '시간': time_str or '',
        '구분': txn.group(1),
        '금액': amount,
        '적요': desc,
        '잔액': balance,
        '계좌': account,
    }, None


def resolve_date(epoch_dt: dt.datetime, month: int, day: int) -> dt.date:
    """문자 원문의 월/일과, 문자를 받은 시각(epoch)을 조합해 정확한 연도를 정한다.

    (문자 자체에는 연도가 없고, 자정을 넘겨 배달되는 경우도 있어 받은 시각과
    가장 가까운 연도를 고른다.)
    """
    candidates = []
    for y in (epoch_dt.year - 1, epoch_dt.year, epoch_dt.year + 1):
        try:
            d = dt.date(y, month, day)
        except ValueError:
            continue
        candidates.append((abs((d - epoch_dt.date()).days), d))
    if not candidates:
        return epoch_dt.date()
    candidates.sort(key=lambda x: x[0])
    return candidates[0][1]


def epoch_ms_to_kst(date_ms: str) -> dt.datetime:
    return dt.datetime.fromtimestamp(int(date_ms) / 1000, tz=KST)


def parse_xml_file(path: str, target_year: int, target_month: int):
    """SMS 백업 XML 하나에서 지정한 연/월의 은행 거래만 뽑아낸다."""
    source = os.path.basename(path)
    records = []
    unrecognized = []

    tree = ET.parse(path)
    root = tree.getroot()

    for sms in root.iter('sms'):
        raw_body = sms.get('body') or ''
        if not any(b in raw_body for b in ('우리', '하나', '농협')):
            continue
        body = extract_body_text(raw_body)
        date_ms = sms.get('date')
        if not date_ms:
            continue
        try:
            epoch_dt = epoch_ms_to_kst(date_ms)
        except (TypeError, ValueError, OSError):
            continue

        for bank_token, _, block in split_messages(body):
            fields, reason = extract_fields(bank_token, block)
            if fields is None:
                unrecognized.append({
                    '은행': BANK_NAMES.get(bank_token, bank_token),
                    '사유': reason,
                    '출처파일': source,
                    '원문': block.strip(),
                })
                continue

            date_val = resolve_date(epoch_dt, fields.pop('month'), fields.pop('day'))
            if date_val.year != target_year or date_val.month != target_month:
                continue

            record = dict(fields)
            record['날짜'] = date_val
            record['확인필요'] = ''
            record['출처파일'] = source
            record['원문'] = block.strip()
            record['정렬시각'] = epoch_dt.replace(tzinfo=None)
            records.append(record)

    return records, unrecognized


# ----------------------------------------------------------------------
# 엑셀 저장
# ----------------------------------------------------------------------

HEADERS = ['은행', '날짜', '시간', '구분', '금액', '적요', '잔액', '계좌', '확인필요', '출처파일', '원문', '정렬시각']
UNRECOG_HEADERS = ['은행', '사유', '출처파일', '원문']


def get_excel_path() -> str:
    folder = os.path.join(os.path.expanduser('~'), 'Documents', '가족_입출금정리')
    os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, '거래내역.xlsx')


def row_key(row: dict):
    return (row['은행'], row['날짜'], row['시간'], row['구분'], row['금액'], row['잔액'])


def sort_key(row: dict):
    """정렬 기준 시각. 문자 원문의 '시:분'은 같은 분에 여러 건이 몰리면 순서를
    구분 못하므로, 문자를 받은 정확한 시각(정렬시각)이 있으면 그걸 우선 쓴다."""
    precise = row.get('정렬시각')
    if isinstance(precise, dt.datetime):
        return precise
    t = row['시간'] or ''
    m = re.match(r'^(\d{1,2}):(\d{2})$', t)
    if m:
        return dt.datetime.combine(row['날짜'], dt.time(int(m.group(1)), int(m.group(2))))
    return dt.datetime.combine(row['날짜'], dt.time(0, 0))


def load_existing_rows(wb) -> list:
    if '전체내역' not in wb.sheetnames:
        return []
    ws = wb['전체내역']
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        row = dict(zip(HEADERS, r))
        if isinstance(row['날짜'], dt.datetime):
            row['날짜'] = row['날짜'].date()
        rows.append(row)
    return rows


def load_existing_unrecognized(wb) -> list:
    if '미인식' not in wb.sheetnames:
        return []
    ws = wb['미인식']
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        rows.append(dict(zip(UNRECOG_HEADERS, r)))
    return rows


def _reorder_by_balance(group_rows: list) -> list:
    """시각(분 단위, 혹은 지연 배달된 문자의 수신 시각)만으로는 거래 순서를
    정확히 알 수 없는 경우가 있다. 잔액은 순서에 따라 유일하게 결정되므로,
    '이전 잔액에서 이 거래를 빼거나 더하면 이 거래의 잔액이 나오는지'를
    근거로 다음 거래를 찾아 순서를 재구성한다. 계좌 하나당 거래 건수가
    많지 않아 전체를 훑어도 부담이 없다."""
    remaining = sorted(group_rows, key=sort_key)
    ordered = []
    prev_balance = None
    while remaining:
        if prev_balance is None:
            ordered.append(remaining.pop(0))
            prev_balance = ordered[-1]['잔액']
            continue
        match_idx = None
        for i, r in enumerate(remaining):
            delta = r['금액'] if r['구분'] == '입금' else -r['금액']
            if prev_balance + delta == r['잔액']:
                match_idx = i
                break
        nxt = remaining.pop(match_idx if match_idx is not None else 0)
        ordered.append(nxt)
        prev_balance = nxt['잔액']
    return ordered


def apply_reconciliation(rows: list):
    groups = {}
    for row in rows:
        key = (row['은행'], row['계좌']) if row['계좌'] else (row['은행'], '')
        groups.setdefault(key, []).append(row)

    for key, group_rows in groups.items():
        ordered = _reorder_by_balance(group_rows)
        prev_balance = None
        for row in ordered:
            if prev_balance is None:
                row['확인필요'] = ''
            else:
                delta = row['금액'] if row['구분'] == '입금' else -row['금액']
                expected = prev_balance + delta
                row['확인필요'] = '' if expected == row['잔액'] else '예'
            prev_balance = row['잔액']


def _reset_sheet(wb, name):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def write_all_sheet(wb, rows):
    ws = _reset_sheet(wb, '전체내역')
    ws.append(HEADERS)
    for row in sorted(rows, key=lambda r: (sort_key(r), r['은행'])):
        ws.append([row.get(h) for h in HEADERS])
    _format_amount_columns(ws, ['금액', '잔액'])


def write_bank_sheets(wb, rows):
    for bank in BANK_ORDER:
        ws = _reset_sheet(wb, bank)
        ws.append(HEADERS)
        bank_rows = [r for r in rows if r['은행'] == bank]
        for row in sorted(bank_rows, key=sort_key):
            ws.append([row.get(h) for h in HEADERS])
        _format_amount_columns(ws, ['금액', '잔액'])


def write_daily_summary(wb, rows):
    ws = _reset_sheet(wb, '일별합계')
    header = ['날짜']
    for bank in BANK_ORDER:
        header += [f'{bank}_입금', f'{bank}_출금']
    header += ['합계_입금', '합계_출금']
    ws.append(header)

    totals = {}
    for row in rows:
        d = row['날짜']
        totals.setdefault(d, {b: {'입금': 0, '출금': 0} for b in BANK_ORDER})
        totals[d][row['은행']][row['구분']] += row['금액']

    for d in sorted(totals.keys()):
        line = [d]
        sum_in = sum_out = 0
        for bank in BANK_ORDER:
            v_in = totals[d][bank]['입금']
            v_out = totals[d][bank]['출금']
            sum_in += v_in
            sum_out += v_out
            line += [v_in, v_out]
        line += [sum_in, sum_out]
        ws.append(line)
    _format_amount_columns(ws, header[1:])


def write_desc_summary(wb, rows):
    ws = _reset_sheet(wb, '적요별합계')
    ws.append(['은행', '적요', '입금합계', '입금건수', '출금합계', '출금건수'])

    agg = {}
    for row in rows:
        desc = row['적요'] or '(미확인)'
        key = (row['은행'], desc)
        agg.setdefault(key, {'입금합계': 0, '입금건수': 0, '출금합계': 0, '출금건수': 0})
        if row['구분'] == '입금':
            agg[key]['입금합계'] += row['금액']
            agg[key]['입금건수'] += 1
        else:
            agg[key]['출금합계'] += row['금액']
            agg[key]['출금건수'] += 1

    items = sorted(agg.items(), key=lambda kv: (kv[0][0], -(kv[1]['입금합계'] + kv[1]['출금합계'])))
    for (bank, desc), v in items:
        ws.append([bank, desc, v['입금합계'], v['입금건수'], v['출금합계'], v['출금건수']])
    _format_amount_columns(ws, ['입금합계', '출금합계'])


def write_unrecognized_sheet(wb, rows):
    ws = _reset_sheet(wb, '미인식')
    ws.append(UNRECOG_HEADERS)
    for row in rows:
        ws.append([row[h] for h in UNRECOG_HEADERS])


def _format_amount_columns(ws, col_names):
    header = [c.value for c in ws[1]]
    for name in col_names:
        if name not in header:
            continue
        idx = header.index(name) + 1
        for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
            for cell in row:
                cell.number_format = '#,##0'


def save_all(new_records: list, new_unrecognized: list):
    path = get_excel_path()
    if os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    existing_rows = load_existing_rows(wb)
    existing_keys = {row_key(r) for r in existing_rows}

    added = 0
    all_rows = existing_rows
    for rec in new_records:
        k = row_key(rec)
        if k in existing_keys:
            continue
        existing_keys.add(k)
        all_rows.append(rec)
        added += 1

    all_unrecognized = load_existing_unrecognized(wb) + new_unrecognized

    apply_reconciliation(all_rows)
    write_all_sheet(wb, all_rows)
    write_bank_sheets(wb, all_rows)
    write_daily_summary(wb, all_rows)
    write_desc_summary(wb, all_rows)
    write_unrecognized_sheet(wb, all_unrecognized)

    wb.save(path)
    review_count = sum(1 for r in all_rows if r['확인필요'] == '예')
    return added, review_count, path


# ----------------------------------------------------------------------
# GUI
# ----------------------------------------------------------------------

class App:
    def __init__(self, root):
        self.root = root
        root.title('가족 입출금 내역 정리')
        root.geometry('600x460')

        self.selected_files = []
        today = dt.date.today()

        tk.Label(root, text='문자 백업 XML 파일(SMS Backup & Restore 등)을 선택하고,\n정리할 연/월을 입력해서 처리하세요.',
                 font=('맑은 고딕', 11), justify='left').pack(pady=(14, 4))

        period_frame = tk.Frame(root)
        period_frame.pack(pady=4)
        tk.Label(period_frame, text='연도:').grid(row=0, column=0, padx=4)
        self.year_var = tk.StringVar(value=str(today.year))
        tk.Entry(period_frame, textvariable=self.year_var, width=6).grid(row=0, column=1, padx=4)
        tk.Label(period_frame, text='월:').grid(row=0, column=2, padx=4)
        self.month_var = tk.StringVar(value=str(today.month))
        tk.Entry(period_frame, textvariable=self.month_var, width=4).grid(row=0, column=3, padx=4)

        btn_frame = tk.Frame(root)
        btn_frame.pack(pady=8)

        self.select_btn = tk.Button(btn_frame, text='XML 파일 선택', width=16,
                                     command=self.on_select)
        self.select_btn.grid(row=0, column=0, padx=6)

        self.process_btn = tk.Button(btn_frame, text='문자 저장하기', width=16,
                                      command=self.on_process, state='disabled')
        self.process_btn.grid(row=0, column=1, padx=6)

        self.open_btn = tk.Button(btn_frame, text='엑셀 열기', width=16,
                                   command=self.on_open_excel)
        self.open_btn.grid(row=0, column=2, padx=6)

        self.selected_label = tk.Label(root, text='선택된 파일: 0개')
        self.selected_label.pack(pady=(4, 8))

        self.log = scrolledtext.ScrolledText(root, height=16, width=72, state='disabled')
        self.log.pack(padx=10, pady=6, fill='both', expand=True)

        self.log_line('"XML 파일 선택" 버튼을 눌러 시작하세요.')

    def log_line(self, text):
        self.log.configure(state='normal')
        self.log.insert('end', text + '\n')
        self.log.see('end')
        self.log.configure(state='disabled')

    def on_select(self):
        files = filedialog.askopenfilenames(
            title='문자 백업 XML 파일 선택',
            filetypes=[('XML 파일', '*.xml'), ('모든 파일', '*.*')],
        )
        if not files:
            return
        self.selected_files = list(files)
        self.selected_label.config(text=f'선택된 파일: {len(self.selected_files)}개')
        self.process_btn.config(state='normal')
        self.log_line(f'{len(self.selected_files)}개 파일을 선택했습니다.')

    def on_open_excel(self):
        path = get_excel_path()
        if not os.path.exists(path):
            messagebox.showinfo('안내', '아직 저장된 엑셀 파일이 없습니다. 먼저 문자를 저장해주세요.')
            return
        os.startfile(path)

    def on_process(self):
        if not self.selected_files:
            return
        try:
            year = int(self.year_var.get())
            month = int(self.month_var.get())
            if not (1 <= month <= 12):
                raise ValueError
        except ValueError:
            messagebox.showerror('오류', '연도/월을 올바르게 입력해주세요. (예: 2026 / 7)')
            return

        self.select_btn.config(state='disabled')
        self.process_btn.config(state='disabled')
        self.log_line(f'{year}년 {month}월 내역을 처리합니다...')
        thread = threading.Thread(target=self._process_worker, args=(year, month), daemon=True)
        thread.start()

    def _process_worker(self, year, month):
        try:
            files = self.selected_files
            all_records = []
            all_unrecognized = []
            for path in files:
                name = os.path.basename(path)
                try:
                    records, unrecognized = parse_xml_file(path, year, month)
                except Exception as e:
                    self.root.after(0, self.log_line, f'[{name}] 읽기 실패: {e}')
                    continue
                all_records.extend(records)
                all_unrecognized.extend(unrecognized)
                self.root.after(0, self.log_line,
                                 f'[{name}] {year}년 {month}월 거래 {len(records)}건, 미인식 {len(unrecognized)}건')

            added, review_count, path = save_all(all_records, all_unrecognized)
            self.root.after(0, self._on_done, added, review_count, len(all_unrecognized))
        except Exception:
            err = traceback.format_exc()
            self.root.after(0, self._on_error, err)

    def _on_done(self, added, review_count, unrecog_count):
        self.log_line(f'완료: 신규 {added}건 저장, 확인 필요 {review_count}건, 미인식 {unrecog_count}건')
        self.select_btn.config(state='normal')
        self.process_btn.config(state='normal')
        messagebox.showinfo(
            '처리 완료',
            f'신규 거래 {added}건을 저장했습니다.\n'
            f'확인 필요(잔액 불일치 의심): {review_count}건\n'
            f'미인식(수동 확인 필요): {unrecog_count}건\n\n'
            f'엑셀 파일: {get_excel_path()}'
        )

    def _on_error(self, err):
        self.log_line('오류가 발생했습니다:')
        self.log_line(err)
        self.select_btn.config(state='normal')
        self.process_btn.config(state='normal')
        messagebox.showerror('오류', '처리 중 오류가 발생했습니다. 로그를 확인해주세요.')


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == '__main__':
    main()
